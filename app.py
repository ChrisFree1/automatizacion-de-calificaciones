import tkinter as tk
from tkinter import filedialog, ttk, messagebox
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook

# Variables para almacenar los nombres de estudiantes de ambos archivos
nombres_profesora = []
nombres_institucional = []

# Variables para almacenar las columnas de calificaciones
columnas_calificaciones_profesora = ["Primer Parcial", "Segundo Parcial", "Nota Final"]
columnas_calificaciones_institucional = []

ubicacion_archivo_institucional = ""

# Definir df_institucional y df_profesora como variables globales
df_institucional = pd.DataFrame()
df_profesora = pd.DataFrame()
archivos_cargados = False

# ...

def capitalizar_nombre(nombre):
    # Divide el nombre en palabras
    palabras = nombre.split()
    # Capitaliza la primera letra de cada palabra
    palabras_capitalizadas = [palabra.capitalize() for palabra in palabras]
    # Une las palabras capitalizadas
    return ' '.join(palabras_capitalizadas)

# ...


def cargar_archivo_profesora():
    global nombres_profesora, columnas_calificaciones_profesora, df_profesora, archivos_cargados
    archivo = filedialog.askopenfilename(filetypes=[("Archivos de Excel", "*.xlsx")])
    if archivo:
        # Leer el archivo Excel de la profesora y capitalizar los nombres
        df_profesora = pd.read_excel(archivo)
        df_profesora['Nombre'] = df_profesora['Nombre'].apply(capitalizar_nombre)
        nombres_profesora = [nombre.lower() for nombre in df_profesora["Nombre"]]
        # Al cargar el archivo de la profesora, habilitar el botón "Subir Excel Institucional"
        columnas_calificaciones_profesora = df_profesora.columns
        subir_institucional_button["state"] = "normal"
        archivos_cargados = True


def cargar_archivo_institucional():
    global nombres_institucional, columnas_calificaciones_institucional, df_institucional, archivos_cargados, ubicacion_archivo_institucional

    if not archivos_cargados:
        messagebox.showerror("Error", "Cargue primero el archivo de la profesora.")
        return

    mensaje_advertencia = "En el Excel que subió anteriormente, los nombres de sus estudiantes deben estar en el Excel institucional para poder transferir las calificaciones. \nCaso contrario no se procederá a transferir."
    mensaje_adicional = ""
    mensaje_recordatorio = "¡Recuerde! Los nombres de sus estudiantes deben ser los mismos que están en el Excel Institucional."

    messagebox.showinfo("Dato Importante", mensaje_advertencia, icon='info', default='ok')

    archivo = filedialog.askopenfilename(filetypes=[("Archivos de Excel", "*.xlsx")])
    if archivo:
        global nombres_institucional, columnas_calificaciones_institucional, ubicacion_archivo_institucional
        nombres_institucional = []  # Restablecer la lista de nombres
        ubicacion_archivo_institucional = archivo  # Guardar la ubicación del archivo institucional
        wb = openpyxl.load_workbook(archivo)
        sheet = wb.active

        # Obtener los nombres capitalizados de la plantilla institucional
        for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1):
            for cell in row:
                nombres_institucional.append(cell.value.lower())

        # Al cargar el archivo institucional, habilitar el botón "Transferir Calificaciones" y verificar las columnas
        columnas_calificaciones_institucional = [cell.value for cell in sheet[1] if cell.value is not None]
        if not list(columnas_calificaciones_profesora) == list(columnas_calificaciones_institucional):
            mensaje_adicional = "Las siguientes columnas de calificaciones no coinciden con el Excel institucional:\n"
            columnas_no_coincidentes = set(columnas_calificaciones_profesora).difference(columnas_calificaciones_institucional)
            mensaje_adicional += f"- {', '.join(columnas_no_coincidentes)}\nPor favor, verifique los nombres de las columnas."
            ventana_no_encontrados(mensaje_adicional, columnas_no_coincidentes)
        else:
            transferir_calificaciones_button["state"] = "normal"


def ventana_no_encontrados(mensaje, columnas_no_coincidentes):
    def on_cierre():
        deshabilitar_botones()
        ventana_emergente.destroy()

    ventana_emergente = tk.Toplevel(ventana)
    ventana_emergente.title("Advertencia")

    # Calcula el ancho y alto de la pantalla
    ancho_pantalla = ventana_emergente.winfo_screenwidth()
    alto_pantalla = ventana_emergente.winfo_screenheight()

    # Calcula las coordenadas para centrar la ventana emergente
    x = (ancho_pantalla - 400) // 2  # 400 es el ancho de tu ventana emergente
    y = (alto_pantalla - 300) // 2   # 200 es la altura de tu ventana emergente

    # Calcula la altura de la ventana emergente en función del número de columnas no coincidentes
    altura_ventana = max(200, 100 + len(columnas_no_coincidentes) * 20)

    # Establece las dimensiones y la posición de la ventana emergente
    ventana_emergente.geometry(f"400x{altura_ventana}+{x}+{y}")
    ventana_emergente.protocol("WM_DELETE_WINDOW", on_cierre)

    # Configurar un estilo para la ventana emergente
    style = ttk.Style()
    style.configure("TLabel", font=("Helvetica", 14))

    # Etiqueta para mostrar el mensaje de advertencia
    label = ttk.Label(ventana_emergente, text=mensaje, foreground="black", wraplength=380)
    label.pack(pady=20)



def transferir_calificaciones():
    global nombres_profesora, nombres_institucional, ubicacion_archivo_institucional

    if not nombres_profesora or not nombres_institucional:
        messagebox.showerror("Error", "Primero cargue los archivos de la profesora y el institucional.")
        return

    if not list(columnas_calificaciones_profesora) == list(columnas_calificaciones_institucional):
        mensaje_adicional = "Las siguientes columnas de calificaciones no coinciden con el Excel institucional:\n"
        columnas_no_coincidentes = set(columnas_calificaciones_profesora).difference(columnas_calificaciones_institucional)
        mensaje_adicional += f"- {', '.join(columnas_no_coincidentes)}\nPor favor, verifique los nombres de las columnas."
        ventana_no_encontrados(mensaje_adicional, columnas_no_coincidentes)
    else:
        # Capitalizar los nombres en df_profesora
        df_profesora['Nombre'] = df_profesora['Nombre'].apply(capitalizar_nombre)
        nombres_profesora = [nombre.lower() for nombre in df_profesora['Nombre']]

        # Transferir calificaciones a la plantilla institucional
        wb = openpyxl.load_workbook(ubicacion_archivo_institucional)
        sheet = wb.active

        for columna in columnas_calificaciones_profesora:
            for nombre_estudiante, calificacion in zip(df_profesora['Nombre'], df_profesora[columna]):
                if nombre_estudiante.lower() in nombres_institucional:
                    # Buscar la fila correspondiente al estudiante
                    row_idx = nombres_institucional.index(nombre_estudiante.lower()) + 2  # +2 para tener en cuenta el encabezado
                    cell = sheet.cell(row=row_idx, column=columnas_calificaciones_institucional.index(columna) + 1)  # +1 para ajustar índice
                    cell.value = calificacion

        wb.save(ubicacion_archivo_institucional)
        mensaje_exitoso = "Las calificaciones han sido transferidas exitosamente."
        messagebox.showinfo("Transferencia Exitosa", mensaje_exitoso)
        deshabilitar_botones()







# Función para deshabilitar los botones y mostrar el mensaje
def deshabilitar_botones():
    subir_profesora_button["state"] = "disabled"
    subir_institucional_button["state"] = "disabled"
    transferir_calificaciones_button["state"] = "disabled"
    mensaje_final.config(text="La transferencia de calificaciones ha sido completada.\nPor favor, verifique y cierre la aplicación.")

# --------------------------------------- Interfaces de mi aplicación -----------------------------------

# Crear la ventana principal
ventana = tk.Tk()
ventana.title("Transferencia de Calificaciones")

# Calcula el ancho y alto de la pantalla
ancho_pantalla = ventana.winfo_screenwidth()
alto_pantalla = ventana.winfo_screenheight()

# Calcula las coordenadas para centrar la ventana principal
x = (ancho_pantalla - 400) // 2  # 400 es el ancho de tu ventana principal
y = (alto_pantalla - 200) // 2   # 200 es la altura de tu ventana principal

# Establece las dimensiones y la posición de la ventana principal
ventana.geometry(f"400x200+{x}+{y}")

# Configurar un estilo para hacer que la GUI sea más moderna
style = ttk.Style()
style.configure("TButton", font=("Helvetica", 12))
style.configure("TLabel", font=("Helvetica", 14))

# Restaurar la apariencia de los mensajes a negro
style.configure("TLabel", foreground="black")

# Botón para cargar el archivo Excel de la profesora
subir_profesora_button = ttk.Button(ventana, text="Subir Excel Profesora", command=cargar_archivo_profesora)
subir_profesora_button.pack(pady=20)

# Botón para cargar el archivo Excel institucional (inicialmente deshabilitado)
subir_institucional_button = ttk.Button(ventana, text="Subir Excel Institucional", command=cargar_archivo_institucional, state="disabled")
subir_institucional_button.pack()

# Botón para transferir calificaciones (inicialmente deshabilitado)
transferir_calificaciones_button = ttk.Button(ventana, text="Transferir Calificaciones", command=transferir_calificaciones, state="disabled")
transferir_calificaciones_button.pack()

# Etiqueta para mostrar el resultado o la alerta
mensaje_final = ttk.Label(ventana, text="", foreground="black")
mensaje_final.pack()

# Personalizar la apariencia de la ventana principal
ventana.resizable(False, False)  # Ventana no redimensionable

# Iniciar la aplicación
ventana.mainloop()
